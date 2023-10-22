import pandas as pd
import openpyxl

books_excel=openpyxl.load_workbook("Books.xlsx")
students=openpyxl.load_workbook("students.xlsx")

book_sheet=books_excel['Sheet1']
students_sheet=students['Sheet1']
# students_pandas=pd.read_excel("students.xlsx")
books_pandas=pd.read_excel("Books.xlsx")

def sign_up(student_row):
    reg=int(input("Enter 1 to do registration else put any number: "))
    if reg!=1:
        print('Thanks for visiting')
        print("***************************")
    else:
        print("**********Registration***********")
        name=str(input("Enter your name: "))
        roll_no=str(input("Enter your roll number: "))
        email=str(input("Enter your email address: "))
        students_sheet['A{}'.format(student_row)]=name
        students_sheet['B{}'.format(student_row)]=roll_no
        students_sheet['C{}'.format(student_row)]=email

        # student_row=student_row+1
        print("you have been registered now you have access to library system")
        print("*********************************************************")
        students.save("students.xlsx")
        get_book()
        return student_row

        
def get_book():
    get_or_not=int(input("press 1 to get book else press any number to avoid: "))
    if get_or_not!=1:
        print("Thanks for visiting")
        print("********************************************************")
    else:
        students_pandas=pd.read_excel("students.xlsx")
        books_pandas=pd.read_excel("Books.xlsx")
        roll_no=str(input("Enter your roll number: ").strip())
        roll_numbers=[str(roll) for roll in students_pandas['roll_no']]
        # print(roll_numbers)

        if roll_no in roll_numbers:
            books_not_available=[]
            books_available=[]
            books=[str(b) for b in books_pandas['book']]
            status=[str(s) for s in books_pandas['status']]

            for i in range(len(status)):
                if status[i] == "yes":
                    books_available.append(books[i])
                else:
                    books_not_available.append(books[i])
            print("Available books: ",books_available)
            print()
            # print(books_not_available)

            book_search=str(input('enter your book name to search: ').strip())
            if book_search in books_available:
                # print(students_pandas[students_pandas['roll_no'].astype(str) == roll_no].values.tolist())
                result_list = students_pandas[students_pandas['roll_no'].astype(str) == roll_no].values.tolist()
                new_result=result_list[0]
                book_index=0
                for i in range (len(book_sheet['A'])):
                    if book_sheet['A{}'.format(i+1)].value==book_search:
                        book_index=i+1
                        book_sheet['B{}'.format(book_index)]="no"
                        book_sheet['C{}'.format(book_index)]=new_result[0]
                        book_sheet['D{}'.format(book_index)]=new_result[1]
                        book_sheet['E{}'.format(book_index)]=new_result[2]
                        books_excel.save('Books.xlsx')
                        break
                print("Book issued")
                print("********************************************************")
            elif book_search in books_not_available:
                for i in range (len(book_sheet['A'])):
                    if book_sheet['A{}'.format(i+1)].value==book_search:
                        book_index=i+1
                        print("Book issued to: ")
                        print("Book: ",book_search,", Status: ",book_sheet['B{}'.format(book_index)].value,
                        ", Name: ",book_sheet['C{}'.format(book_index)].value,
                        ", Roll_no: ",book_sheet['D{}'.format(book_index)].value,
                        ", Email: ",book_sheet['E{}'.format(book_index)].value)
                        print("***************************************************")
                        break
                        # books_excel.save('Books.xlsx')
                # print("That has been issued to other student")
            else:
                print("That book is not available in Libraray") 
                print("*********************************************************")   

        else:
            print("***********you are not registered")
            print("**********************************")
            student_row = (students_sheet.max_row)+1
            student_row1=sign_up(student_row)
            student_row=student_row+student_row1
            students.save('students.xlsx')

def return_book():
    students_pandas=pd.read_excel("students.xlsx")
    roll_no=str(input("Enter your roll number: ").strip())
    books_pandas=pd.read_excel("Books.xlsx")
    roll_numbers=[str(roll) for roll in students_pandas['roll_no']]
    if roll_no in roll_numbers:
        books_not_available=[]
        books_available=[]
        books=[str(b) for b in books_pandas['book']]
        status=[str(s) for s in books_pandas['status']]

        for i in range(len(status)):
            if status[i] == "yes":
                books_available.append(books[i])
            else:
                books_not_available.append(books[i])
        # print(books_available)
        # print(books_not_available)

        book_search=str(input('enter your book name to search: ').strip())
        if book_search in books_not_available:
            book_index=1
            for i in range(len(book_sheet['A'])):
                if book_sheet['A{}'.format(i+1)].value == book_search:
                    if str(book_sheet['D{}'.format(i+1)].value) == roll_no:
                        book_index=i+1
                        # print("done")
                        book_sheet['B{}'.format(book_index)]="yes"
                        book_sheet['C{}'.format(book_index)]=""
                        book_sheet['D{}'.format(book_index)]=""
                        book_sheet['E{}'.format(book_index)]=""
                        books_excel.save('Books.xlsx')
                        print("Book has been returned")
                        print("***************************************************")
                        break
                    else:
                        print("Wrong credentials")
                        print("***********************************************")
                        break
        elif book_search in books_available:
            print("Enter correct book name")
            print('***************************************')
        else:
            print("Enter correct book name")
            print('***************************************')
    else:
        print("************You are not registered*************")
        sign_up_for_register=str(input("Enter yes to sign up else press anything: "))
        if sign_up_for_register=="yes":
            student_row = (students_sheet.max_row)+1
            student_row1=sign_up(student_row)
            student_row=student_row+student_row1
            students.save('students.xlsx')

def detail_about_book_stocks():
    books_pandas=pd.read_excel("Books.xlsx")
    books_not_available=[]
    books_available=[]
    books=[str(b) for b in books_pandas['book']]
    status=[str(s) for s in books_pandas['status']]

    for i in range(len(status)):
        if status[i] == "yes":
            books_available.append(books[i])
        else:
            books_not_available.append(books[i])
    print("Total books are: ",len(books))
    print("******************************")
    print("Books are: ")
    print(books)
    print("*******************************")
    print("total available books are: ",len(books_available))
    print("Books that are available: ")
    print(books_available)
    print("************************************")
    print("total issued books are: ",len(books_not_available))
    print("Books that has been issued : ")
    print(books_not_available)
    print("***************************************")
    

while True:
    print("*******************************")
    print('press 1 to get book from library ')
    print('press 2 to return book into library')
    print('press 3 to get detail about library books')
    print('press 4 to exit')
    print("*****************************************")
    try:
        choice=int(input('enter what you want: '))
        if choice==1:
            get_book()
        elif choice==2:
            return_book()
        elif choice==3:
            detail_about_book_stocks()
        elif choice==4:
            print("Thanks for visiting")
            print("*********************")
            break
        else:
            print("Invalid Input")
            print("*************************")
    except Exception as e:
        print(e)
        print("invalid input")
        print("***********************")
        



